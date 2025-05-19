import fitz  # PyMuPDF
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from tkinter import messagebox, END

def remove_all_annotations(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        for page_num in range(len(doc)):
            page = doc[page_num]
            annotations = page.annots()
            if annotations:
                for annot in annotations:
                    page.delete_annot(annot)
        doc.saveIncr()  # Save the changes to the same file
        print(f"All annotations removed from: {pdf_path}")
    except Exception as e:
        print(f"Error processing {pdf_path}: {str(e)}")

def extract_words_from_box(page, rect):
    words = page.get_text("words")
    return [word for word in words if rect.intersects(fitz.Rect(word[:4]))]

def clean_price_format(price_str):
    price_str = price_str.replace('.', '').replace(',', '.')
    try:
        return float(price_str)
    except ValueError:
        return None

def extract_item_info(doc, page_num, item_rect, box_definitions, po_number, destination, item_number):
    page = doc[page_num]
    item_info = {'PO number': po_number, 'Destination': destination, 'Item number': item_number}
    for box_name, (x0, y0_offset, x1, y1_offset) in box_definitions.items():
        rect = fitz.Rect(x0, item_rect.y0 + y0_offset, x1, item_rect.y0 + y1_offset)
        words_in_box = extract_words_from_box(page, rect)
        
        if words_in_box:
            words_in_box.sort(key=lambda word: (word[3], word[0]))
            item_info[box_name] = ' '.join(word[4] for word in words_in_box)
        else:
            item_info[box_name] = "No text found"
    
    if 'material_revision_box' in item_info:
        material_revision_text = item_info['material_revision_box']
        revision_match = re.search(r"Material Revision:\s*([\w-]+)", material_revision_text)
        if revision_match:
            item_info['Revision'] = revision_match.group(1)
        else:
            # Check the next page for material revision if not found on the current page
            if page_num + 1 < len(doc):
                next_page = doc[page_num + 1]
                next_page_rect = fitz.Rect(80, 98, 192, 154)  # coordinates for the next page material revision box
                next_page_words = extract_words_from_box(next_page, next_page_rect)
                if next_page_words:
                    next_page_words.sort(key=lambda word: (word[3], word[0]))
                    next_page_text = ' '.join(word[4] for word in next_page_words)
                    revision_match_next = re.search(r"Material Revision:\s*([\w-]+)", next_page_text)
                    item_info['Revision'] = revision_match_next.group(1) if revision_match_next else "No revision found"
                else:
                    item_info['Revision'] = "No revision found"
            else:
                item_info['Revision'] = "No revision found"
    
    if 'net_price' in item_info:
        item_info['net_price'] = clean_price_format(item_info['net_price'])

    return item_info

def determine_destination(destination_text):
    destination_text_lower = destination_text.lower()
    if "india" in destination_text_lower:
        return "India"
    elif "china" in destination_text_lower:
        return "China"
    elif "hankkion" in destination_text_lower:
        return "Etu"
    elif "crusher" in destination_text_lower:
        return "Mob"
    elif "scm" in destination_text_lower:
        return "SCM"
    elif "holtum noordweg 5" in destination_text_lower:
        return "HO5"
    elif "holtum noordweg 11" in destination_text_lower:
        return "HO11"
    elif "helsinki" in destination_text_lower:
        return "VA"
    elif "huko" in destination_text_lower:
        return "Lokomo"
    elif "urusvuorenkatu 5" in destination_text_lower:
        return "Turku"         
    else:
        return "Unknown"

def convert_date_format(date_str):
    try:
        return datetime.strptime(date_str, '%Y%b%d').strftime('%d.%m.%Y')
    except ValueError:
        return date_str

def print_text_in_boxes(pdf_path):
    item_number_box_def = (40, 85, 70, 705)
    po_number_box_def = (330, 40, 570, 115)
    destination_box_def = (40, 200, 330, 300)
    other_boxes_definitions = {
        'product_code': (72, -2, 131, 15),
        'due_date': (217, -2, 270, 15),
        'qty': (286, -2, 314, 15),
        'net_price': (350, -2, 399, 15),
        'material_revision_box': (80, 15, 192, 80)
    }

    po_number = None
    destination = "Unknown"
    extracted_data = []

    try:
        doc = fitz.open(pdf_path)
        for page_num in range(len(doc)):
            page = doc[page_num]
            if page_num == 0:
                po_number_rect = fitz.Rect(*po_number_box_def)
                words_in_po_number_box = extract_words_from_box(page, po_number_rect)
                po_numbers = [word[4] for word in words_in_po_number_box if word[4].isdigit() and len(word[4]) == 10]
                if po_numbers:
                    po_number = po_numbers[0]

                destination_rect = fitz.Rect(*destination_box_def)
                words_in_destination_box = extract_words_from_box(page, destination_rect)
                destination_text = ' '.join(word[4] for word in words_in_destination_box)
                destination = determine_destination(destination_text)

            item_number_rect = fitz.Rect(*item_number_box_def)
            words_in_item_number_box = extract_words_from_box(page, item_number_rect)
            
            item_numbers = [word for word in words_in_item_number_box if word[4].isdigit() and len(word[4]) == 5 and word[4][0] == '0' and int(word[4]) % 10 == 0]
            for item_number in item_numbers:
                item_number_value = str(int(item_number[4]))  # Remove leading zeros
                item_number_rect = fitz.Rect(item_number[:4])
                item_info = extract_item_info(doc, page_num, item_number_rect, other_boxes_definitions, po_number, destination, item_number_value)
                
                if 'product_code' in item_info:
                    description_parts = item_info['product_code'].split(' ', 1)
                    if len(description_parts) == 2:
                        item_info['product_code'] = description_parts[0]
                        item_info['product_name'] = description_parts[1]
                    else:
                        item_info['product_code'] = description_parts[0]
                        item_info['product_name'] = 'No text found'

                if 'due_date' in item_info:
                    item_info['due_date'] = convert_date_format(item_info['due_date'])

                extracted_data.append(item_info)

    except Exception as e:
        print(f"Error processing {pdf_path}: {str(e)}")

    return extracted_data

def find_last_non_empty_row(sheet):
    last_row = sheet.max_row
    while last_row > 0:
        if any(cell.value is not None for cell in sheet[last_row]):
            break
        last_row -= 1
    return last_row

def start_extraction(pdf_file_paths, excel_file_path, log_text):
    if not pdf_file_paths:
        messagebox.showinfo("No selection", "No PDF files selected. Exiting.")
        return

    if not excel_file_path:
        messagebox.showinfo("No selection", "No Excel file selected. Exiting.")
        return

    all_data = []
    po_numbers = []

    try:
        log_text.configure(state='normal')
        log_text.delete(1.0, END)
        log_text.insert(END, "Attempting to open PDF files:\n")
        log_text.insert(END, "Removing annotations:\n")
        log_text.insert(END, "Processing:\n")

        for pdf_path in pdf_file_paths:
            # Remove all annotations before processing
            remove_all_annotations(pdf_path)

            extracted_data = print_text_in_boxes(pdf_path)
            if extracted_data:
                po_numbers.append(extracted_data[0]['PO number'])
            all_data.extend(extracted_data)


        df = pd.DataFrame(all_data)    

        cols = df.columns.tolist()
        destination_index = cols.index('Destination')
        cols.insert(destination_index + 1, cols.pop(cols.index('Item number')))
        

        # Directly select and order the columns as needed
        df = df[['Destination', 'PO number', 'Item number', 'product_code', 'Revision', 'product_name', 'due_date', 'qty', 'net_price']]
        df.columns = ['Destination:', 'PO number:', 'Pos:', 'Product code:', 'Revision:', 'Product name:', 'Due date:', 'Qty:', 'Net price:']

        # Convert to appropriate types
        df['PO number:'] = pd.to_numeric(df['PO number:'], errors='coerce')
        df['Pos:'] = pd.to_numeric(df['Pos:'], errors='coerce')
        df['Qty:'] = pd.to_numeric(df['Qty:'], errors='coerce')
        df['Net price:'] = pd.to_numeric(df['Net price:'], errors='coerce').round(2)

        if df.empty:
            messagebox.showinfo("Error", "No data extracted from the PDFs.")
            log_text.insert(END, "No data extracted.\n")
            log_text.configure(state='disabled')
            return

        log_text.insert(END, "Data extracted:\n")

        book = load_workbook(excel_file_path)
        
        # Ensure the sheet "TILAUKSET" exists
        if "TILAUKSET" not in book.sheetnames:
            messagebox.showinfo("Error", "Sheet 'TILAUKSET' not found in the Excel file.")
            log_text.insert(END, "Sheet 'TILAUKSET' not found.\n")
            log_text.configure(state='disabled')
            return
        
        sheet = book["TILAUKSET"]
        last_row = find_last_non_empty_row(sheet)
        start_row = last_row + 1

        for r_idx, row in df.iterrows():
            for c_idx, value in enumerate(row):
                sheet.cell(row=start_row + r_idx, column=c_idx + 1, value=value)

        book.save(excel_file_path)

        log_text.insert(END, f"Data successfully appended to the Excel file:\n{excel_file_path}\n")
        log_text.insert(END, "\nProcessed PO Numbers:\n")
        for po in po_numbers:
            log_text.insert(END, po + "\n")

        log_text.configure(state='disabled')

        messagebox.showinfo("Success", "Data successfully appended to the Excel file.")

    except Exception as e:
        messagebox.showerror("Error", str(e))
        log_text.insert(END, f"Error: {str(e)}\n")
        log_text.configure(state='disabled')
