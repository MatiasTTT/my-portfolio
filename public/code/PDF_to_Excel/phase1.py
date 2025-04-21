import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import re
from tkinter import END

def phase1_validation(excel_file_path, validation_log):
    try:
        # Load the workbook and the necessary sheets
        wb = load_workbook(excel_file_path)
        ws_kuormat = wb['KUORMAT']
        ws_p4t = wb['P4T']

        # Convert the data from the sheets into pandas DataFrames
        df_kuormat = pd.DataFrame(ws_kuormat.values)
        df_p4t = pd.DataFrame(ws_p4t.values)

        # Assuming the first row is the header
        df_kuormat.columns = df_kuormat.iloc[0]
        df_kuormat = df_kuormat[1:].reset_index(drop=True)
        df_p4t.columns = df_p4t.iloc[0]
        df_p4t = df_p4t[1:].reset_index(drop=True)

        # Print first few rows of the P4T dataframe for verification
        print("First few rows of P4T dataframe:")
        print(df_p4t.head())

        # UUSI Strip exclamation marks from 'Del. Qty:' column in the DataFrame
        df_kuormat['Del. Qty:'] = df_kuormat['Del. Qty:'].astype(str).str.replace('!', '')

        # UUSI Write the cleaned 'Del. Qty:' column back to the worksheet
        for idx, value in enumerate(df_kuormat['Del. Qty:']):
            if value.strip().isdigit():  # Ensure the value is numeric
                ws_kuormat.cell(row=idx+2, column=7, value=float(value))
            else:
                ws_kuormat.cell(row=idx+2, column=7, value=None)  # Leave empty if not numeric

        # Ensure numeric columns are correctly formatted
        df_kuormat['PO nr.:'] = pd.to_numeric(df_kuormat['PO nr.:'], errors='coerce').astype('Int64')
        df_kuormat['Del. Qty:'] = pd.to_numeric(df_kuormat['Del. Qty:'], errors='coerce')
        df_kuormat['Total'] = pd.to_numeric(df_kuormat['Total'], errors='coerce')
        df_p4t['Contract no.'] = pd.to_numeric(df_p4t['Contract no.'], errors='coerce')
        df_p4t['Shipped qty'] = pd.to_numeric(df_p4t['Shipped qty'], errors='coerce')

        # Ensure the 'Code:' column is treated as a string and trim spaces
        df_kuormat['Code:'] = df_kuormat['Code:'].astype(str).str.strip()
        df_p4t['Material no.'] = df_p4t['Material no.'].astype(str).str.strip()

        # Function to strip revision indicators
        def strip_revision(code):
            # Remove any part of the code after a space or dash
            return re.split(r'[\s-]', code)[0]

        # Track rows with total price as 0
        zero_price_rows = []
        # Track rows that failed validation
        failed_validation_rows = []

        # Iterate through each row in KUORMAT
        for index, row in df_kuormat.iterrows():
            po_nr = row['PO nr.:']
            code = row['Code:']
            del_qty = row['Del. Qty:']
            total = row['Total']
            if pd.isna(po_nr) or pd.isna(code) or pd.isna(del_qty):
                print(f"Skipping row {index+2}: Invalid data found.")
                continue

            # Strip the revision from the code
            code_base = strip_revision(code)
            print(f"Processing row {index+2}: PO nr.: {po_nr}, Code: {code}, Base Code: {code_base}, Del. Qty: {del_qty}")

            # Find matching rows in P4T
            match = df_p4t[
                (df_p4t['Contract no.'] == po_nr) &
                (df_p4t['Material no.'].str.startswith(code_base)) &
                (df_p4t['Shipped qty'] == del_qty)
            ]

            if match.empty:
                print(f"No match found for row {index+2}: PO nr.: {po_nr}, Base Code: {code_base}, Del. Qty: {del_qty}")
                print(f"Relevant rows in P4T for PO nr. {po_nr} and Base Code {code_base}:")
                print(df_p4t[(df_p4t['Contract no.'] == po_nr)])
                if len(str(po_nr)) == 10:  # Ensure the PO number is 10 digits
                    failed_validation_rows.append((int(po_nr), code))  # Ensure po_nr is an integer
            else:
                print(f"Match found for row {index+2}: PO nr.: {po_nr}, Base Code: {code_base}, Del. Qty: {del_qty}")
                # If a match is found, highlight the row in KUORMAT
                for cell in ws_kuormat.iter_rows(min_row=index + 2, max_row=index + 2, min_col=1, max_col=14):
                    for c in cell:
                        c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                # Check if the total is 0 after validation
                if total == 0 and len(str(po_nr)) == 10:
                    zero_price_rows.append((int(po_nr), code))  # Ensure po_nr is an integer
                    for cell in ws_kuormat.iter_rows(min_row=index + 2, max_row=index + 2, min_col=14, max_col=14):  # Column N is the 14th column
                        for c in cell:
                            c.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            c.font = Font(color="9C0006")

        # Save the workbook
        wb.save(excel_file_path)
        print("Phase 1 validation completed successfully.")

        # Report zero price rows and failed validation rows in the GUI
        validation_log.configure(state='normal')
        if zero_price_rows:
            validation_log.insert(END, "Price detected as 0,00 for the following rows:\n")
            for po_nr, code in zero_price_rows:
                validation_log.insert(END, f"-PO: {po_nr}. Code: {code}.\n", "zero_price")
        if failed_validation_rows:
            validation_log.insert(END, "Validation was unsuccessful for the following rows:\n")
            for po_nr, code in failed_validation_rows:
                validation_log.insert(END, f"-PO: {po_nr}. Code: {code}.\n", "failed_validation")
        validation_log.insert(END, "\nValidation completed.\n")
        validation_log.tag_config("zero_price", foreground="#7984f7")
        validation_log.tag_config("failed_validation", foreground="#fc4949")
        validation_log.configure(state='disabled')

    except Exception as e:
        print(f"An error occurred during Phase 1 validation: {str(e)}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        phase1_validation(sys.argv[1])
    else:
        print("Please provide the path to the Excel file.")
