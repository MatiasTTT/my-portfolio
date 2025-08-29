import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
from tkinter import Text

def phase2_update(excel_file_path):
    fully_delivered_rows = []
    partially_delivered_rows = []
    df_tilaukset = None
    
    try:
        # Load the workbook and the necessary sheets
        wb = load_workbook(excel_file_path)
        ws_tilaukset = wb['TILAUKSET']
        ws_kuormat = wb['KUORMAT']
        ws_p4t = wb['P4T']

        # Convert the data from the sheets into pandas DataFrames
        df_tilaukset = pd.DataFrame(ws_tilaukset.values)
        df_kuormat = pd.DataFrame(ws_kuormat.values)
        df_p4t = pd.DataFrame(ws_p4t.values)

        # Assuming the first row is the header
        df_tilaukset.columns = df_tilaukset.iloc[0]
        df_tilaukset = df_tilaukset[1:].reset_index(drop=True)
        df_kuormat.columns = df_kuormat.iloc[0]
        df_kuormat = df_kuormat[1:].reset_index(drop=True)
        df_p4t.columns = df_p4t.iloc[0]
        df_p4t = df_p4t[1:].reset_index(drop=True)

        # Ensure numeric columns are correctly formatted
        df_tilaukset['PO number:'] = pd.to_numeric(df_tilaukset['PO number:'], errors='coerce')
        df_tilaukset['Pos:'] = pd.to_numeric(df_tilaukset['Pos:'], errors='coerce')
        df_tilaukset['Qty:'] = pd.to_numeric(df_tilaukset['Qty:'], errors='coerce')
        df_tilaukset['Delivered:'] = pd.to_numeric(df_tilaukset['Delivered:'], errors='coerce').fillna(0)
        df_tilaukset['Remaining:'] = pd.to_numeric(df_tilaukset['Remaining:'], errors='coerce').fillna(df_tilaukset['Qty:'])

        df_kuormat['PO nr.:'] = pd.to_numeric(df_kuormat['PO nr.:'], errors='coerce')
        df_kuormat['Order Qty:'] = pd.to_numeric(df_kuormat['Order Qty:'], errors='coerce')
        df_kuormat['Del. Qty:'] = pd.to_numeric(df_kuormat['Del. Qty:'], errors='coerce')

        df_p4t['PO position no.'] = pd.to_numeric(df_p4t['PO position no.'], errors='coerce')
        df_p4t['Shipped qty'] = pd.to_numeric(df_p4t['Shipped qty'], errors='coerce')

        # Function to strip revision indicators and handle numeric codes
        def strip_revision(code):
            if pd.isna(code):
                return None, None
            code_str = str(code).strip()
            parts = code_str.split(' ', 1)
            return parts[0], parts[1] if len(parts) > 1 else None

        # To keep track of rows that are matched and updated
        updated_indices = set()

        # Iterate through each row in TILAUKSET
        for index, row in df_tilaukset.iterrows():
            po_number = row['PO number:']
            pos = row['Pos:']
            product_code = row['Product code:']
            revision = row['Revision:']
            qty = row['Qty:']
            delivered = row['Delivered:']

            if pd.isna(po_number) or pd.isna(pos) or pd.isna(product_code):
                print(f"Skipping row {index + 2}: Invalid data found.")
                continue

            print(f"Processing row {index + 2}: PO number: {po_number}, Pos: {pos}, Product code: {product_code}, Revision: {revision}, Qty: {qty}, Delivered: {delivered}")

            # Match with KUORMAT
            product_code_base, product_revision = strip_revision(product_code)
            if product_code_base is None:
                print(f"Skipping row {index + 2}: Invalid product code.")
                continue
            
            if product_revision is None:
                product_revision = ''
            revision = str(revision).lstrip('0').replace('-', '')  # Remove dash and leading zeros from revision

            print(f"Stripped product code: {product_code_base}, Stripped revision: {revision}")

            # Improved matching logic
            kuormat_match = df_kuormat[
                (df_kuormat['PO nr.:'] == po_number) & 
                (df_kuormat['Code:'].apply(lambda x: str(x).startswith(str(product_code_base))))
            ]

            if kuormat_match.empty:
                print(f"No match found in KUORMAT for PO number: {po_number}, Product code: {product_code}")
                continue

            for k_index, k_row in kuormat_match.iterrows():
                k_product_code_base, k_product_revision = strip_revision(k_row['Code:'])
                print(f"Checking KUORMAT row {k_index + 2}: PO nr.: {k_row['PO nr.:']}, Code: {k_row['Code:']}, Order Qty: {k_row['Order Qty:']}, Del. Qty: {k_row['Del. Qty:']}")

                # Skip if the revisions do not match, ignoring dashes
                k_product_revision = str(k_product_revision).replace('-', '') if k_product_revision else ''
                if k_product_revision and k_product_revision != revision:
                    print(f"Skipping KUORMAT row {k_index + 2} due to revision mismatch: {k_product_revision} != {revision}")
                    continue

                print(f"Matched KUORMAT row {k_index + 2}: PO nr.: {k_row['PO nr.:']}, Code: {k_row['Code:']}, Order Qty: {k_row['Order Qty:']}, Del. Qty: {k_row['Del. Qty:']}")

                # Match with P4T
                p4t_match = df_p4t[
                    (df_p4t['Contract no.'] == po_number) & 
                    (df_p4t['PO position no.'] == pos)
                ]

                if p4t_match.empty:
                    print(f"No match found in P4T for PO number: {po_number}, Pos: {pos}")
                    continue

                for p_index, p_row in p4t_match.iterrows():
                    shipped_qty = p_row['Shipped qty']
                    print(f"Matched P4T row {p_index + 2}: Contract no.: {p_row['Contract no.']}, PO position no.: {p_row['PO position no.']}, Shipped qty: {shipped_qty}")

                    # Update Delivered and Remaining
                    new_delivered = delivered + shipped_qty
                    new_remaining = qty - new_delivered

                    df_tilaukset.at[index, 'Delivered:'] = new_delivered
                    df_tilaukset.at[index, 'Remaining:'] = max(new_remaining, 0)
                    updated_indices.add(index)

                    if new_remaining != 0:
                        partially_delivered_rows.append(index + 2)

        # Write back to the TILAUKSET sheet only for updated rows
        for r_idx in updated_indices:
            row = df_tilaukset.iloc[r_idx]
            for c_idx, value in enumerate(row):
                ws_tilaukset.cell(row=r_idx + 2, column=c_idx + 1, value=value)

        # Apply coloring after all updates are done
        fully_delivered_rows = apply_coloring(ws_tilaukset, df_tilaukset, updated_indices)

        wb.save(excel_file_path)
        print("Phase 2 update completed successfully.")
        
    except Exception as e:
        import traceback
        print(f"An error occurred during Phase 2 update: {str(e)}")
        traceback.print_exc()
    
    return fully_delivered_rows, partially_delivered_rows, df_tilaukset



def phase2_update_with_reporting(excel_file_path):
    fully_delivered_rows, partially_delivered_rows, df_tilaukset = phase2_update(excel_file_path)

    fully_delivered_count = len(fully_delivered_rows)
    partially_delivered_count = len(partially_delivered_rows)
    
    # Prepare the new message format
    message_parts = []
    message_tags = []
    
    message_parts.append("Fully delivered rows:\n")
    message_tags.append(None)
    
    message_parts.append(f"-Count: {fully_delivered_count}\n")
    message_tags.append("full_rows")
    
    if partially_delivered_count > 0:
        message_parts.append("Partially delivered rows:\n")
        message_tags.append(None)
        
        message_parts.append(f"-Count: {partially_delivered_count}\n")
        message_tags.append("partial_rows")
        
        for row in partially_delivered_rows:
            po_nr = df_tilaukset.at[row - 2, 'PO number:']
            code = df_tilaukset.at[row - 2, 'Product code:']
            message_parts.append(f"-PO: {po_nr}. Code: {code}.\n")
            message_tags.append("partial_rows")
    
    message_parts.append("Order ledger update completed.\n")
    message_tags.append(None)
    
    return message_parts, message_tags



def apply_coloring(ws_tilaukset, df_tilaukset, updated_indices):
    # Define the fills for the different conditions
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Get the current date in Finnish standard format
    current_date = datetime.now().strftime("%d.%m.%Y")
    italic_font = Font(italic=True)

    fully_delivered_rows = []

    for index in updated_indices:
        row = df_tilaukset.iloc[index]
        qty = row['Qty:']
        delivered = row['Delivered:']
        remaining = row['Remaining:']

        # Apply green fill if fully delivered
        if remaining == 0 and delivered == qty:
            for cell in ws_tilaukset.iter_rows(min_row=index + 2, max_row=index + 2, min_col=1, max_col=11):  # Ensure it covers all columns
                for c in cell:
                    c.fill = green_fill
            fully_delivered_rows.append(index + 2)  # Record as fully delivered

        # Apply orange fill if there are remaining quantities
        if remaining > 0:
            ws_tilaukset.cell(row=index + 2, column=11).fill = orange_fill

        # Apply blue fill if partially delivered
        if delivered != 0 and delivered != qty:
            ws_tilaukset.cell(row=index + 2, column=10).fill = blue_fill

        # Log the current date in the "Updated:" column (Column L) with italic font
        updated_cell = ws_tilaukset.cell(row=index + 2, column=12, value=current_date)
        updated_cell.font = italic_font

    return fully_delivered_rows


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        fully_delivered, partially_delivered = phase2_update(sys.argv[1])
        print(f"Fully delivered rows: {fully_delivered}")
        print(f"Partially delivered rows: {partially_delivered}")
    else:
        print("Please provide the path to the Excel file.")
